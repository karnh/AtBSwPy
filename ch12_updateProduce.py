#! python3
# updateProduce.py - Correct costs in product sales spreasheet
 
import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# Updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# Loop through rows and updated prices
for rowNum in range(2, sheet.max_row + 1):
    produce_name = sheet.cell(row=rowNum, column = 1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produce_name]

wb.save('updatedProduceSales.xlsx')